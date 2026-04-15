from __future__ import annotations

"""
CueBI Tally Connector — parses Tally XML and Excel exports into
queryable PostgreSQL staging tables.

Supported Tally export formats:
1. XML (Data Interchange) — exported via Alt+E from TallyPrime
   Contains TALLYMESSAGE with VOUCHER, LEDGER, STOCKITEM nodes
2. Excel (.xlsx) — exported via Alt+E from TallyPrime reports
   Day Book, Trial Balance, P&L, Balance Sheet, Stock Summary, etc.

Flow:
  User uploads Tally file → CueBI parses it → creates staging tables
  in PostgreSQL → runs embedding pipeline → user can query with NL
"""
import io
import xml.etree.ElementTree as ET
from typing import Any
from dataclasses import dataclass, field


@dataclass
class TallyVoucher:
    """A single Tally voucher (transaction)."""
    date: str
    voucher_type: str  # Sales, Purchase, Payment, Receipt, Journal, Contra
    voucher_number: str = ""
    party_name: str = ""
    amount: float = 0.0
    narration: str = ""
    ledger_entries: list[dict] = field(default_factory=list)
    inventory_entries: list[dict] = field(default_factory=list)


@dataclass
class TallyLedger:
    """A Tally ledger (account head)."""
    name: str
    parent_group: str = ""
    opening_balance: float = 0.0
    closing_balance: float = 0.0
    gstin: str = ""
    address: str = ""
    state: str = ""


@dataclass
class TallyStockItem:
    """A Tally stock/inventory item."""
    name: str
    parent_group: str = ""
    unit: str = ""
    opening_qty: float = 0.0
    opening_value: float = 0.0
    closing_qty: float = 0.0
    closing_value: float = 0.0
    hsn_code: str = ""
    gst_rate: float = 0.0


@dataclass
class TallyData:
    """Parsed Tally data — ready for staging."""
    company_name: str = ""
    vouchers: list[TallyVoucher] = field(default_factory=list)
    ledgers: list[TallyLedger] = field(default_factory=list)
    stock_items: list[TallyStockItem] = field(default_factory=list)
    raw_tables: dict[str, list[dict]] = field(default_factory=dict)  # For Excel imports


def _safe_float(text: str | None) -> float:
    """Convert Tally amount string to float. Handles Dr/Cr markers, commas."""
    if not text:
        return 0.0
    text = text.strip().replace(",", "")
    is_credit = "Cr" in text
    text = text.replace("Dr", "").replace("Cr", "").strip()
    if not text:
        return 0.0
    try:
        value = float(text)
        return -value if is_credit else value
    except ValueError:
        return 0.0


def _get_text(elem: ET.Element | None, tag: str, default: str = "") -> str:
    """Safely get text from an XML element."""
    if elem is None:
        return default
    child = elem.find(tag)
    if child is not None and child.text:
        return child.text.strip()
    # Try case-insensitive (Tally XML uses all caps)
    for c in elem:
        if c.tag.upper() == tag.upper() and c.text:
            return c.text.strip()
    return default


def parse_tally_xml(xml_content: str | bytes) -> TallyData:
    """
    Parse Tally XML export (Data Interchange format).
    Handles: TALLYMESSAGE containing VOUCHER, LEDGER, STOCKITEM nodes.
    """
    if isinstance(xml_content, str):
        xml_content = xml_content.encode("utf-8")

    data = TallyData()

    try:
        root = ET.fromstring(xml_content)
    except ET.ParseError:
        # Try wrapping in root element
        try:
            root = ET.fromstring(b"<ROOT>" + xml_content + b"</ROOT>")
        except ET.ParseError as e:
            raise ValueError(f"Invalid Tally XML: {e}")

    # Find TALLYMESSAGE (could be at root or nested under BODY/DATA)
    tally_msgs = root.findall(".//TALLYMESSAGE")
    if not tally_msgs:
        tally_msgs = [root]  # Root itself might contain vouchers

    # Extract company name
    company = root.find(".//COMPANY") or root.find(".//SVCURRENTCOMPANY")
    if company is not None and company.text:
        data.company_name = company.text.strip()

    for msg in tally_msgs:
        # Parse Vouchers
        for voucher in msg.findall("VOUCHER"):
            v = TallyVoucher(
                date=_get_text(voucher, "DATE"),
                voucher_type=_get_text(voucher, "VOUCHERTYPENAME"),
                voucher_number=_get_text(voucher, "VOUCHERNUMBER"),
                party_name=_get_text(voucher, "PARTYLEDGERNAME"),
                narration=_get_text(voucher, "NARRATION"),
            )

            # Ledger entries (Dr/Cr)
            for entry in voucher.findall(".//ALLLEDGERENTRIES.LIST") + voucher.findall(".//LEDGERENTRIES.LIST"):
                ledger_name = _get_text(entry, "LEDGERNAME")
                amount = _safe_float(_get_text(entry, "AMOUNT"))
                v.ledger_entries.append({
                    "ledger_name": ledger_name,
                    "amount": amount,
                    "is_debit": amount >= 0,
                })
                if abs(amount) > abs(v.amount):
                    v.amount = abs(amount)

            # Inventory entries
            for inv in voucher.findall(".//ALLINVENTORYENTRIES.LIST") + voucher.findall(".//INVENTORYENTRIES.LIST"):
                v.inventory_entries.append({
                    "stock_item": _get_text(inv, "STOCKITEMNAME"),
                    "quantity": _safe_float(_get_text(inv, "ACTUALQTY") or _get_text(inv, "BILLEDQTY")),
                    "rate": _safe_float(_get_text(inv, "RATE")),
                    "amount": _safe_float(_get_text(inv, "AMOUNT")),
                })

            data.vouchers.append(v)

        # Parse Ledgers
        for ledger in msg.findall("LEDGER"):
            data.ledgers.append(TallyLedger(
                name=_get_text(ledger, "NAME") or ledger.attrib.get("NAME", ""),
                parent_group=_get_text(ledger, "PARENT"),
                opening_balance=_safe_float(_get_text(ledger, "OPENINGBALANCE")),
                closing_balance=_safe_float(_get_text(ledger, "CLOSINGBALANCE")),
                gstin=_get_text(ledger, "PARTYGSTIN") or _get_text(ledger, "GSTIN"),
                address=_get_text(ledger, "ADDRESS"),
                state=_get_text(ledger, "LEDSTATENAME") or _get_text(ledger, "COUNTRYOFRESIDENCE"),
            ))

        # Parse Stock Items
        for item in msg.findall("STOCKITEM"):
            data.stock_items.append(TallyStockItem(
                name=_get_text(item, "NAME") or item.attrib.get("NAME", ""),
                parent_group=_get_text(item, "PARENT"),
                unit=_get_text(item, "BASEUNITS"),
                opening_qty=_safe_float(_get_text(item, "OPENINGBALANCE")),
                opening_value=_safe_float(_get_text(item, "OPENINGVALUE")),
                hsn_code=_get_text(item, "HSNCODE") or _get_text(item, "GSTDETAILS.LIST/HSNCODE"),
                gst_rate=_safe_float(_get_text(item, "GSTDETAILS.LIST/TAXABILITY")),
            ))

    return data


def parse_tally_excel(file_bytes: bytes, filename: str = "") -> TallyData:
    """
    Parse Tally Excel export (.xlsx).
    Auto-detects: Day Book, Trial Balance, P&L, Balance Sheet, Stock Summary
    based on sheet names and column headers.
    """
    import openpyxl

    data = TallyData()
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        headers = []

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            # Clean None values
            cleaned = [str(cell).strip() if cell is not None else "" for cell in row]

            # Skip completely empty rows
            if all(c == "" or c == "None" for c in cleaned):
                continue

            # First non-empty row with multiple values = headers
            if not headers and sum(1 for c in cleaned if c and c != "None") >= 2:
                headers = [c.replace("None", "").strip() for c in cleaned]
                continue

            if headers:
                row_dict = {}
                for j, h in enumerate(headers):
                    if h and j < len(cleaned):
                        row_dict[h] = cleaned[j] if cleaned[j] != "None" else ""
                if any(v for v in row_dict.values()):
                    rows.append(row_dict)

        if rows:
            # Normalize sheet name for table naming
            table_name = _normalize_table_name(sheet_name)
            data.raw_tables[table_name] = rows

            # Auto-detect and parse known Tally report types
            _detect_and_parse_report(data, table_name, headers, rows)

    wb.close()
    return data


def _normalize_table_name(name: str) -> str:
    """Convert sheet name to a valid SQL table name."""
    import re
    name = name.strip().lower()
    name = re.sub(r'[^a-z0-9_]', '_', name)
    name = re.sub(r'_+', '_', name).strip('_')
    if not name:
        name = "tally_data"
    return f"tally_{name}"


def _detect_and_parse_report(
    data: TallyData, table_name: str, headers: list[str], rows: list[dict]
) -> None:
    """Auto-detect Tally report type and parse into structured data."""
    header_lower = [h.lower() for h in headers if h]

    # Day Book detection: has Date, Voucher Type, Debit, Credit
    if any("date" in h for h in header_lower) and any("voucher" in h for h in header_lower):
        for row in rows:
            date_val = _find_value(row, ["date"])
            vtype = _find_value(row, ["voucher type", "vch type", "type"])
            party = _find_value(row, ["particulars", "party", "ledger", "party name"])
            debit = _safe_float(_find_value(row, ["debit", "dr", "debit amount"]))
            credit = _safe_float(_find_value(row, ["credit", "cr", "credit amount"]))
            narration = _find_value(row, ["narration", "remarks"])
            vch_no = _find_value(row, ["vch no", "voucher no", "vch no."])

            if date_val or vtype:
                amount = debit if debit > 0 else credit
                data.vouchers.append(TallyVoucher(
                    date=date_val, voucher_type=vtype, voucher_number=vch_no,
                    party_name=party, amount=amount, narration=narration,
                ))

    # Trial Balance / Ledger detection: has Ledger/Particulars, Debit, Credit
    elif any("particular" in h or "ledger" in h for h in header_lower) and \
         any("debit" in h or "closing" in h for h in header_lower):
        for row in rows:
            name = _find_value(row, ["particulars", "ledger", "ledger name", "account"])
            group = _find_value(row, ["group", "parent", "under"])
            debit = _safe_float(_find_value(row, ["debit", "dr"]))
            credit = _safe_float(_find_value(row, ["credit", "cr"]))
            closing = debit - credit if (debit or credit) else 0

            if name:
                data.ledgers.append(TallyLedger(
                    name=name, parent_group=group, closing_balance=closing,
                ))

    # Stock Summary detection: has Item/Stock, Quantity, Value/Rate
    elif any("stock" in h or "item" in h for h in header_lower) and \
         any("qty" in h or "quantity" in h or "value" in h for h in header_lower):
        for row in rows:
            name = _find_value(row, ["stock item", "item", "name", "particulars"])
            qty = _safe_float(_find_value(row, ["quantity", "qty", "closing qty"]))
            value = _safe_float(_find_value(row, ["value", "amount", "closing value"]))
            unit = _find_value(row, ["unit", "uom"])

            if name:
                data.stock_items.append(TallyStockItem(
                    name=name, unit=unit, closing_qty=qty, closing_value=value,
                ))


def _find_value(row: dict, possible_keys: list[str]) -> str:
    """Find a value in a dict by trying multiple possible key names (case-insensitive)."""
    for key in possible_keys:
        for actual_key, value in row.items():
            if key.lower() in actual_key.lower():
                return str(value).strip() if value else ""
    return ""